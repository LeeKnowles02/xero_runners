import os
import re
import json
import logging
import time
import random
import hashlib
from io import BytesIO
from datetime import datetime, timezone, timedelta
from email.utils import format_datetime
from typing import Dict, Any, List, Tuple, Optional, Set

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

XERO_API_BASE = "https://api.xero.com/api.xro/2.0"
logger = logging.getLogger("xero_runner.xero_jobs")

# ---------------------------
# Tuning
# ---------------------------
# Global throttle between ALL API calls (helps minute-limit 429s).
# 1.2s ≈ 50 req/min, leaving room for UI polling and other endpoints.
XERO_THROTTLE_SECONDS = float(os.getenv("XERO_THROTTLE_SECONDS", "1.2"))

# JournalLines checkpoint flush frequency (rows appended between saves)
JOURNALLINES_FLUSH_EVERY = int(os.getenv("JOURNALLINES_FLUSH_EVERY", "500"))

# JournalLines resume checkpoint file name
JOURNALLINES_CHECKPOINT_NAME = "JournalLines.checkpoint.json"

# JournalLines processed-marker sheet
JOURNALLINES_PROCESSED_SHEET = "JournalLines_Processed"

# Journals gap-fill behavior
JOURNALS_STOP_AFTER_NO_NEW_BATCHES = int(os.getenv("JOURNALS_STOP_AFTER_NO_NEW_BATCHES", "3"))

# Journals "created-date overlap" behavior (bulletproof)
JOURNALS_CREATED_BUFFER_DAYS = int(os.getenv("JOURNALS_CREATED_BUFFER_DAYS", "7"))

# Hard cap on offset batches in "gap-fill" mode (safety)
JOURNALS_GAPFILL_MAX_BATCHES = int(os.getenv("JOURNALS_GAPFILL_MAX_BATCHES", "5000"))

# ---------------------------
# Endpoint registry
# ---------------------------
ENDPOINTS: Dict[str, Dict[str, Any]] = {
    "Organisation": {
        "path": "/Organisation",
        "root": "Organisations",
        "paged": False,
        "columns": [
            "OrganisationID", "Name", "BaseCurrency", "CountryCode",
            "Version", "OrganisationEntityType", "FinancialYearEndDay",
            "FinancialYearEndMonth", "PeriodLockDate", "EndOfYearLockDate",
            "CreatedDateUTC", "Timezone"
        ],
    },
    "Journals": {
        "path": "/Journals",
        "root": "Journals",
        "paged": True,
        "columns": ["JournalID", "JournalNumber", "JournalDate", "CreatedDateUTC", "Reference", "SourceID", "SourceType"],
    },
    "JournalLines": {
        "path": "/Journals",
        "root": "JournalLines",
        "paged": False,
        "special": "journal_lines",
        "columns": [
            "JournalID",
            "JournalNumber",
            "JournalDate",
            "SourceType",
            "SourceID",
            "JournalLine.AccountCode",
            "JournalLine.Description",
            "JournalLine.LineAmount",
            "JournalLine.TaxType",
            "JournalLine.TaxAmount",
            "JournalLine.Tracking",
        ],
    },
    "Contacts": {
        "path": "/Contacts",
        "root": "Contacts",
        "paged": False,
        "columns": ["ContactID", "Name", "EmailAddress", "ContactStatus", "IsCustomer", "IsSupplier", "UpdatedDateUTC"],
    },
    "Invoices": {
        "path": "/Invoices",
        "root": "Invoices",
        "paged": True,
        "columns": [
            "InvoiceID", "InvoiceNumber", "Type", "Status", "Date", "DueDate", "UpdatedDateUTC",
            "Contact.ContactID", "Contact.Name",
            "Reference",
            "SubTotal", "TotalTax", "Total", "AmountDue", "AmountPaid", "AmountCredited",
            "CurrencyCode", "ExchangeRate",
        ],
    },
    "Payments": {
        "path": "/Payments",
        "root": "Payments",
        "paged": True,
        "columns": [
            "PaymentID", "Invoice.InvoiceID", "Invoice.InvoiceNumber", "Date", "Amount",
            "Reference", "CurrencyRate", "UpdatedDateUTC",
        ],
    },
    "BankTransactions": {
        "path": "/BankTransactions",
        "root": "BankTransactions",
        "paged": True,
        "columns": [
            "BankTransactionID", "Type", "Status", "Date", "Reference",
            "Contact.ContactID", "Contact.Name",
            "SubTotal", "TotalTax", "Total",
            "CurrencyCode", "ExchangeRate",
            "UpdatedDateUTC",
        ],
    },
    "Accounts": {
        "path": "/Accounts",
        "root": "Accounts",
        "paged": False,
        "columns": [
            "AccountID", "Code", "Name", "Type", "Class", "Status",
            "TaxType", "EnablePaymentsToAccount", "UpdatedDateUTC",
        ],
    },
    "TrackingCategories": {
        "path": "/TrackingCategories",
        "root": "TrackingCategories",
        "paged": False,
        "columns": [
            "TrackingCategoryID", "Name", "Status",
            "Options",
        ],
    },
    "TaxRates": {
        "path": "/TaxRates",
        "root": "TaxRates",
        "paged": False,
        "columns": [
            "TaxType", "Name", "Status", "ReportTaxType", "CanApplyToAssets",
            "CanApplyToEquity", "CanApplyToExpenses", "CanApplyToLiabilities",
            "CanApplyToRevenue",
        ],
    },
}

# ---------------------------
# Date helpers
# ---------------------------
def utc_now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def iso_to_dt(s: str) -> datetime:
    """
    Accepts:
      - 2026-02-23T06:50:41Z
      - 2026-02-23T06:50:41+00:00
      - with fractional seconds in either form
    Returns tz-aware UTC datetime.
    """
    if not s:
        raise ValueError("Empty datetime string")
    ss = str(s).strip()
    if ss.endswith("Z"):
        ss = ss[:-1] + "+00:00"
    dt = datetime.fromisoformat(ss)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    return dt

def parse_xero_date_maybe(val: Any) -> Any:
    if not isinstance(val, str):
        return val
    m = re.match(r"^/Date\((\d+)([+-]\d{4})?\)/$", val)
    if not m:
        return val
    ms = int(m.group(1))
    dt = datetime.fromtimestamp(ms / 1000.0, tz=timezone.utc)
    return dt.strftime("%Y-%m-%dT%H:%M:%SZ")

def _dt_from_created_date_maybe(val: Any) -> Optional[datetime]:
    """
    Robust parse for CreatedDateUTC-like fields:
      - 2024-09-26T06:44:10Z
      - 2024-09-26T06:44:10+00:00
      - may be empty/None
    """
    if val in (None, ""):
        return None
    try:
        return iso_to_dt(str(val))
    except Exception:
        return None

# ---------------------------
# Excel helpers
# ---------------------------
_INVALID_SHEET_CHARS = r'[:\\/?*\[\]]'

def excel_safe_sheet_name(name: str) -> str:
    s = re.sub(_INVALID_SHEET_CHARS, "-", (name or "").strip())
    s = s[:31].rstrip()
    return s or "Sheet"

def excel_unique_sheet_name(wb, desired: str) -> str:
    base = excel_safe_sheet_name(desired)
    if base not in wb.sheetnames:
        return base
    for i in range(2, 1000):
        suffix = f" ({i})"
        cut = 31 - len(suffix)
        candidate = (base[:cut]).rstrip() + suffix
        if candidate not in wb.sheetnames:
            return candidate
    return base[:28] + "_X"

def ensure_excel(path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if os.path.exists(path):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Runs"
    ws.append(["TimestampUTC", "Endpoint", "Mode", "RowsWritten", "Status", "Error"])
    wb.save(path)

def _utc_ts_compact() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

def backup_file(path: str, backup_dir: str, tag: str) -> Optional[str]:
    try:
        if not os.path.exists(path):
            return None
        os.makedirs(backup_dir, exist_ok=True)
        base = os.path.basename(path)
        bname = f"{_utc_ts_compact()}__{tag}__{base}"
        bpath = os.path.join(backup_dir, bname)
        import shutil
        shutil.copy2(path, bpath)
        return bpath
    except Exception as e:
        logger.warning("Backup failed for %s: %s", path, e)
        return None

def save_workbook_atomic(wb, excel_path: str) -> None:
    d = os.path.dirname(excel_path)
    os.makedirs(d, exist_ok=True)
    tmp_path = os.path.join(d, f".tmp_{os.getpid()}_{_utc_ts_compact()}_{os.path.basename(excel_path)}")
    wb.save(tmp_path)
    os.replace(tmp_path, excel_path)

def autosize_columns(ws, max_width: int = 60):
    for col_idx, col in enumerate(ws.columns, start=1):
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, length + 2), max_width)

def append_run_log(wb, endpoint: str, mode: str, rows_written: int, status: str, error: Optional[str]) -> None:
    ws = wb["Runs"]
    ws.append([utc_now_iso(), endpoint, mode, rows_written, status, error or ""])
    autosize_columns(ws)

def list_endpoints() -> List[str]:
    return list(ENDPOINTS.keys())


def workbook_with_only_sheet(excel_path: str, endpoint_name: str) -> Optional[BytesIO]:
    """
    Returns a BytesIO containing an Excel workbook with only the sheet for the given endpoint.
    Returns None if the sheet does not exist.
    """
    if not os.path.isfile(excel_path):
        return None
    wb = load_workbook(excel_path, read_only=False)
    sheet_name = excel_safe_sheet_name(endpoint_name)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    ws_src = wb[sheet_name]
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name
    for row in ws_src.iter_rows(values_only=True):
        new_ws.append(list(row))
    wb.close()
    buf = BytesIO()
    new_wb.save(buf)
    buf.seek(0)
    return buf

def endpoint_columns(endpoint_name: str) -> List[str]:
    if endpoint_name not in ENDPOINTS:
        raise ValueError(f"Unknown endpoint: {endpoint_name}")
    return ENDPOINTS[endpoint_name]["columns"]

# ---- minimal inserts for JournalLines completeness ----
def _ensure_sheet(wb, sheet_name: str, columns: List[str]):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
        return ws
    ws = wb[sheet_name]
    _ensure_sheet_header(ws, columns)
    return ws

def _load_first_col_set(ws) -> Set[str]:
    out: Set[str] = set()
    if ws.max_row < 2:
        return out
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        out.add(str(v))
    return out

# ---------------------------
# Journals / JournalLines linkage helpers
# ---------------------------
def _stable_hash(parts: List[Any]) -> str:
    payload = "\x1f".join("" if p is None else str(p).strip() for p in parts)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()

def _journals_header_hash_from_row_values(row_values: List[Any], columns: List[str]) -> str:
    """
    Compute a stable hash of the Journals header row.
    Uses all columns except JournalID (ID itself doesn't signal change).
    """
    if not row_values or not columns:
        return ""
    include_idxs = [i for i, c in enumerate(columns) if c != "JournalID"]
    parts = [row_values[i] if i < len(row_values) else None for i in include_idxs]
    return _stable_hash(parts)

def _load_journals_header_map(ws_journals, columns: List[str]) -> Dict[str, Dict[str, Any]]:
    """
    Returns: JournalID -> {JournalNumber, CreatedDateUTC, HeaderHash}
    """
    out: Dict[str, Dict[str, Any]] = {}
    if ws_journals.max_row < 2:
        return out
    ncols = len(columns)
    for r in range(2, ws_journals.max_row + 1):
        jid = ws_journals.cell(row=r, column=1).value  # JournalID col1
        if jid is None:
            continue
        vals = [ws_journals.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        out[str(jid)] = {
            "JournalNumber": vals[1] if ncols >= 2 else None,
            "CreatedDateUTC": vals[3] if ncols >= 4 else None,
            "HeaderHash": _journals_header_hash_from_row_values(vals, columns),
        }
    return out

def _max_created_date_from_journals_sheet(ws_journals, journals_cols: List[str]) -> Optional[datetime]:
    """
    Scans CreatedDateUTC column in Journals sheet; returns max datetime (UTC).
    """
    if ws_journals.max_row < 2:
        return None
    try:
        idx_created = journals_cols.index("CreatedDateUTC") + 1
    except ValueError:
        return None
    max_dt: Optional[datetime] = None
    for r in range(2, ws_journals.max_row + 1):
        v = ws_journals.cell(row=r, column=idx_created).value
        dt = _dt_from_created_date_maybe(v)
        if dt is None:
            continue
        if max_dt is None or dt > max_dt:
            max_dt = dt
    return max_dt

def _journal_number_set_from_sheet(ws_journals, journals_cols: List[str]) -> Set[int]:
    """
    Returns set of JournalNumber ints present in sheet (best-effort).
    """
    out: Set[int] = set()
    if ws_journals.max_row < 2:
        return out
    try:
        idx_num = journals_cols.index("JournalNumber") + 1
    except ValueError:
        return out
    for r in range(2, ws_journals.max_row + 1):
        v = ws_journals.cell(row=r, column=idx_num).value
        if v is None or v == "":
            continue
        try:
            out.add(int(v))
        except Exception:
            continue
    return out

def _detect_discontinuity(journal_numbers: Set[int]) -> bool:
    """
    Heuristic: if there are big gaps in JournalNumber range, signal discontinuity.
    We consider any missing in the last 1000 range as a sign; cheap and safe.
    """
    if not journal_numbers:
        return False
    mx = max(journal_numbers)
    mn = min(journal_numbers)
    if mx - mn <= 1000:
        return False
    tail_min = mx - 1000
    tail = [n for n in journal_numbers if n >= tail_min]
    if len(tail) < 900:
        return True
    return False

# ---------------------------
# JSON helpers
# ---------------------------
def get_by_path(obj: Dict[str, Any], path: str) -> Any:
    cur: Any = obj
    for part in path.split("."):
        if cur is None:
            return None
        if isinstance(cur, dict):
            cur = cur.get(part)
        else:
            return None
    cur = parse_xero_date_maybe(cur)
    if isinstance(cur, (list, dict)):
        return str(cur)
    return cur

def _ensure_sheet_header(ws, columns: List[str]) -> None:
    if ws.max_row == 0:
        ws.append(columns)
        return
    if ws.max_row >= 1:
        existing = [c.value for c in ws[1]]
        if existing != columns:
            ws.delete_rows(1, ws.max_row)
            ws.append(columns)

def write_sheet_selected_columns(wb, sheet_name: str, items: List[Dict[str, Any]], columns: List[str]) -> int:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    if not items:
        ws.append(["(no rows returned)"])
        autosize_columns(ws)
        return 0

    ws.append(columns)
    for item in items:
        ws.append([get_by_path(item, col) for col in columns])

    autosize_columns(ws)
    return len(items)

def merge_sheet_selected_columns(
    wb,
    sheet_name: str,
    items: List[Dict[str, Any]],
    columns: List[str],
    key_col: str
) -> int:
    """
    Incremental-safe: merges items into an existing sheet keyed by key_col.
    - Updates existing rows with same key
    - Appends new keys
    """
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
        for item in items:
            ws.append([get_by_path(item, col) for col in columns])
        autosize_columns(ws)
        return len(items)

    ws = wb[sheet_name]
    _ensure_sheet_header(ws, columns)

    key_idx = columns.index(key_col) + 1
    key_to_row: Dict[str, int] = {}

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=key_idx).value
        if v is None:
            continue
        key_to_row[str(v)] = r

    updated = 0
    appended = 0

    for item in items:
        k = get_by_path(item, key_col)
        if k is None:
            continue
        k = str(k)
        row_values = [get_by_path(item, col) for col in columns]
        if k in key_to_row:
            r = key_to_row[k]
            for c, val in enumerate(row_values, start=1):
                ws.cell(row=r, column=c).value = val
            updated += 1
        else:
            ws.append(row_values)
            appended += 1

    autosize_columns(ws)
    return updated + appended

# ---------------------------
# JournalLines row dedupe
# ---------------------------
def _norm_cell(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    return s.strip()

def _row_key_from_values(values: List[Any]) -> str:
    payload = "\x1f".join(_norm_cell(v) for v in values)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()

def _load_existing_row_keys(ws, columns: List[str]) -> Set[str]:
    keys: Set[str] = set()
    if ws.max_row < 2:
        return keys
    ncols = len(columns)
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        keys.add(_row_key_from_values(vals))
    return keys

# ---------------------------
# HTTP helpers
# ---------------------------
def _add_if_modified_since(headers: Dict[str, str], watermark_iso: str) -> Dict[str, str]:
    out = dict(headers)
    dt = iso_to_dt(watermark_iso)
    out["If-Modified-Since"] = format_datetime(dt)
    return out


def _fetch_generic_endpoint(
    endpoint_name: str,
    headers: Dict[str, str],
    incremental_since_iso: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Fetch items for any endpoint that has path/root (no special logic).
    Returns list of item dicts. Uses If-Modified-Since when incremental_since_iso is set.
    """
    cfg = ENDPOINTS[endpoint_name]
    path = cfg["path"]
    root = cfg["root"]
    paged = cfg.get("paged", False)
    url = f"{XERO_API_BASE}{path}"
    use_headers = _add_if_modified_since(headers, incremental_since_iso) if incremental_since_iso else dict(headers)
    all_items: List[Dict[str, Any]] = []

    if paged:
        page = 1
        while page <= 1000:
            r = _get_json(url, headers=use_headers, params={"page": page})
            if r.status_code == 304:
                return []
            j = r.json()
            batch = j.get(root, []) if isinstance(j, dict) else []
            if not batch:
                break
            all_items.extend(batch)
            if len(batch) < 100:
                break
            page += 1
    else:
        r = _get_json(url, headers=use_headers)
        if r.status_code == 304:
            return []
        j = r.json()
        raw = j.get(root, []) if isinstance(j, dict) else []
        if isinstance(raw, list):
            all_items = raw
        elif raw is not None:
            all_items = [raw]
        else:
            all_items = []

    return all_items

def _refresh_headers_after_401(current_headers: Dict[str, str]) -> Optional[Dict[str, str]]:
    tenant_id = current_headers.get("xero-tenant-id")
    if not tenant_id:
        return None
    try:
        from xero_auth import headers_from_token_file
        return headers_from_token_file(tenant_id)
    except Exception as e:
        logger.warning("401 recovery: could not refresh headers from token file (%s)", e)
        return None

def _force_refresh_headers(current_headers: Dict[str, str]) -> Optional[Dict[str, str]]:
    tenant_id = current_headers.get("xero-tenant-id")
    if not tenant_id:
        return None
    try:
        from xero_auth import refresh_token_file_inplace, headers_from_token_file, default_token_path
        cid = os.environ.get("XERO_CLIENT_ID")
        csec = os.environ.get("XERO_CLIENT_SECRET")
        if not cid or not csec:
            logger.warning("401 force-refresh skipped (missing XERO_CLIENT_ID/SECRET in env)")
            return None
        token_path = default_token_path()
        refresh_token_file_inplace(token_path, cid, csec)
        return headers_from_token_file(tenant_id)
    except Exception as e:
        logger.warning("401 force-refresh failed (%s)", e)
        return None

def _get_json(url: str, headers: Dict[str, str], params: Optional[Dict[str, Any]] = None) -> requests.Response:
    max_attempts = 8
    base_sleep = 1.0

    for attempt in range(1, max_attempts + 1):
        time.sleep(XERO_THROTTLE_SECONDS)

        r = requests.get(url, headers=headers, params=params or {}, timeout=90)

        if r.status_code == 304:
            return r

        if r.status_code == 401:
            new_headers = _refresh_headers_after_401(headers)
            if new_headers:
                logger.warning("Xero 401: reloaded Authorization from token file; retrying url=%s", url)
                headers.update(new_headers)
                time.sleep(XERO_THROTTLE_SECONDS)
                r = requests.get(url, headers=headers, params=params or {}, timeout=90)
                if r.status_code == 304:
                    return r
                if r.status_code == 200:
                    return r

            forced = _force_refresh_headers(headers)
            if forced:
                logger.warning("Xero 401: forced token refresh; retrying url=%s", url)
                headers.update(forced)
                time.sleep(XERO_THROTTLE_SECONDS)
                r = requests.get(url, headers=headers, params=params or {}, timeout=90)
                if r.status_code == 304:
                    return r
                if r.status_code == 200:
                    return r

            r.raise_for_status()
            return r

        if r.status_code != 429:
            r.raise_for_status()
            return r

        retry_after = r.headers.get("Retry-After")
        rate_problem = (
            r.headers.get("X-Rate-Limit-Problem")
            or r.headers.get("X-Rate-Limit-Reason")
            or "unknown"
        )
        min_left = r.headers.get("X-MinLimit-Remaining")
        day_left = r.headers.get("X-DayLimit-Remaining")

        # Daily quota exhausted: don't wait 50+ minutes, fail fast with a clear message
        if str(day_left) == "0" and (retry_after is None or float(retry_after) > 300):
            raise RuntimeError(
                "Xero daily API rate limit is exhausted (day_left=0). "
                "Wait until the limit resets (usually midnight UTC) or reduce the number of API calls. url=%s" % url
            )

        if retry_after is not None:
            sleep_s = float(retry_after)
        else:
            sleep_s = min(60.0, base_sleep * (2 ** (attempt - 1)))
            sleep_s += random.uniform(0, 0.5)

        # Cap sleep so we don't block for 50+ minutes (Xero can send Retry-After ~3000s)
        MAX_429_SLEEP = 120
        if sleep_s > MAX_429_SLEEP:
            logger.warning("Capping 429 sleep from %.0fs to %ds (day_left=%s)", sleep_s, MAX_429_SLEEP, day_left)
            sleep_s = MAX_429_SLEEP

        logger.warning(
            "Xero 429 (problem=%s min_left=%s day_left=%s) sleeping %.2fs (attempt %s/%s) url=%s",
            rate_problem, min_left, day_left, sleep_s, attempt, max_attempts, url
        )
        time.sleep(sleep_s)

    raise RuntimeError(
        "Xero rate limit exceeded (429). If day_left=0, your daily API quota is used. "
        "Wait until tomorrow or check your Xero app rate limits. url=%s" % url
    )

# ---------------------------
# Journals: correct pagination using offset (supports overlap window and gap fill)
# ---------------------------
def _fetch_journals_offset(
    headers: Dict[str, str],
    incremental_since_iso: Optional[str],
    batch_limit: int = 5000,
    existing_ids: Optional[Set[str]] = None,
    stop_after_no_new_batches: Optional[int] = None,
    created_cutoff_dt: Optional[datetime] = None,
) -> Tuple[List[Dict[str, Any]], bool]:
    use_headers = headers
    if incremental_since_iso:
        use_headers = _add_if_modified_since(headers, incremental_since_iso)

    url = f"{XERO_API_BASE}/Journals"
    out_items: List[Dict[str, Any]] = []

    offset: Optional[int] = None
    last_offset: Optional[int] = None
    no_new_streak = 0
    stop_after = stop_after_no_new_batches if stop_after_no_new_batches is not None else None
    reached_cutoff = False
    old_cutoff_streak = 0

    for _ in range(batch_limit):
        params: Dict[str, Any] = {}
        if offset is not None:
            params["offset"] = offset

        r = _get_json(url, headers=use_headers, params=params)
        if r.status_code == 304:
            return [], True

        j = r.json()
        if isinstance(j, dict) and "Journals" not in j:
            raise ValueError(f"Expected root key 'Journals' but got keys {list(j.keys())} for url {url}")

        batch = j.get("Journals", [])
        if not batch:
            break

        if created_cutoff_dt is not None:
            all_older = True
            for it in batch:
                dt = _dt_from_created_date_maybe(it.get("CreatedDateUTC"))
                if dt is None:
                    all_older = False
                    continue
                if dt >= created_cutoff_dt:
                    all_older = False
            if all_older:
                old_cutoff_streak += 1
            else:
                old_cutoff_streak = 0
            if old_cutoff_streak >= 3:
                reached_cutoff = True
                break

        new_in_batch = 0

        for it in batch:
            jid = it.get("JournalID")
            if not jid:
                continue
            sjid = str(jid)

            if existing_ids is not None and sjid in existing_ids:
                continue

            if created_cutoff_dt is not None:
                dt = _dt_from_created_date_maybe(it.get("CreatedDateUTC"))
                if dt is not None and dt < created_cutoff_dt:
                    continue

            out_items.append(it)
            new_in_batch += 1

            if existing_ids is not None:
                existing_ids.add(sjid)

        if stop_after is not None:
            if new_in_batch == 0:
                no_new_streak += 1
                if no_new_streak >= stop_after:
                    break
            else:
                no_new_streak = 0

        if len(batch) < 100:
            break

        next_offset = batch[-1].get("JournalNumber")
        if next_offset is None:
            break

        if next_offset == last_offset:
            logger.warning("Journals offset did not advance (offset=%s). Stopping to avoid repeats.", next_offset)
            break

        last_offset = next_offset
        offset = next_offset

    return out_items, reached_cutoff

# ---------------------------
# JournalLines: checkpoint helpers
# ---------------------------
def _checkpoint_path_for_excel(excel_path: str) -> str:
    d = os.path.dirname(excel_path)
    cdir = os.path.join(d, "checkpoints")
    os.makedirs(cdir, exist_ok=True)
    return os.path.join(cdir, JOURNALLINES_CHECKPOINT_NAME)

def _load_checkpoint(path: str) -> Optional[Dict[str, Any]]:
    try:
        if not os.path.exists(path):
            return None
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def _save_checkpoint(path: str, data: Dict[str, Any]) -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    os.replace(tmp, path)

def _delete_checkpoint(path: str) -> None:
    try:
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass

# ---------------------------
# JournalLines: rewrite helper
# ---------------------------
def _rewrite_journallines_sheet_excluding(
    wb,
    sheet_name: str,
    columns: List[str],
    exclude_journal_ids: Set[str]
) -> None:
    kept_rows: List[List[Any]] = []
    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        for r in range(2, ws_old.max_row + 1):
            jid = ws_old.cell(row=r, column=1).value
            if jid is None:
                continue
            if str(jid) in exclude_journal_ids:
                continue
            kept_rows.append([ws_old.cell(row=r, column=c).value for c in range(1, len(columns) + 1)])
        del wb[sheet_name]

    ws_new = wb.create_sheet(sheet_name)
    ws_new.append(columns)
    for row in kept_rows:
        ws_new.append(row)

# ---------------------------
# Main runner
# ---------------------------
def run_endpoint_selected(
    endpoint_name: str,
    headers: Dict[str, str],
    excel_path: str,
    selected_columns: Optional[List[str]] = None,
    incremental_since_iso: Optional[str] = None,
) -> Tuple[int, str, str, Optional[str]]:
    """
    Returns: (rows_written, status, mode, error)
    """
    ensure_excel(excel_path)
    wb = load_workbook(excel_path)

    try:
        allowed = set(endpoint_columns(endpoint_name))
        if not selected_columns:
            selected_columns = endpoint_columns(endpoint_name)

        bad = [c for c in selected_columns if c not in allowed]
        if bad:
            raise ValueError(f"Invalid columns for {endpoint_name}: {bad}")

        assert "xero-tenant-id" in headers and headers["xero-tenant-id"], "Missing tenant id"

        # -----------------------
        # Journals: bulletproof run
        # -----------------------
        if endpoint_name == "Journals":
            sheet_name = "Journals"
            cols = ENDPOINTS["Journals"]["columns"]

            if sheet_name not in wb.sheetnames:
                wsj = wb.create_sheet(sheet_name)
                wsj.append(cols)
            else:
                wsj = wb[sheet_name]
                _ensure_sheet_header(wsj, cols)

            max_created = _max_created_date_from_journals_sheet(wsj, cols)
            if max_created is None:
                created_cutoff = None
                mode = "CANONICAL BASELINE (no cutoff)"
            else:
                created_cutoff = max_created - timedelta(days=JOURNALS_CREATED_BUFFER_DAYS)
                mode = f"OVERLAP_CREATED >= {created_cutoff.strftime('%Y-%m-%dT%H:%M:%SZ')} (buffer={JOURNALS_CREATED_BUFFER_DAYS}d)"

            items_a, _ = _fetch_journals_offset(
                headers=headers,
                incremental_since_iso=None,
                batch_limit=50 * 20,
                existing_ids=None,
                stop_after_no_new_batches=None,
                created_cutoff_dt=created_cutoff,
            )
            rows_a = merge_sheet_selected_columns(wb, sheet_name, items_a, cols, "JournalID")

            numbers = _journal_number_set_from_sheet(wsj, cols)
            discontinuity = _detect_discontinuity(numbers)

            rows_b = 0
            if discontinuity:
                existing_ids2: Set[str] = _load_first_col_set(wsj)
                items_b, _ = _fetch_journals_offset(
                    headers=headers,
                    incremental_since_iso=None,
                    batch_limit=JOURNALS_GAPFILL_MAX_BATCHES,
                    existing_ids=existing_ids2,
                    stop_after_no_new_batches=None,
                    created_cutoff_dt=None,
                )
                rows_b = merge_sheet_selected_columns(wb, sheet_name, items_b, cols, "JournalID")

            rows_written = rows_a + rows_b
            mode2 = mode + (f" + GAPFILL({rows_b})" if discontinuity else "")
            append_run_log(wb, endpoint_name, mode2, rows_written, "OK", None)
            save_workbook_atomic(wb, excel_path)
            logger.info("Wrote %s rows to %s (mode=%s)", rows_written, endpoint_name, mode2)
            return rows_written, "OK", mode2, None

        # -----------------------
        # JournalLines: self-heal checkpoint + missing/changed
        # -----------------------
        if endpoint_name == "JournalLines":
            sheet_name = "JournalLines"
            checkpoint_path = _checkpoint_path_for_excel(excel_path)
            ck = _load_checkpoint(checkpoint_path)

            # ✅ Self-heal: if checkpoint exists but has no pending_ids, rebuild plan
            if ck and ck.get("endpoint") == "JournalLines" and ck.get("excel_path") == excel_path:
                pending = ck.get("pending_ids") or []
                if not pending:
                    logger.warning("Checkpoint has pending_ids=0; rebuilding JournalLines plan from workbook state.")
                    _delete_checkpoint(checkpoint_path)
                    ck = None

            if "Journals" not in wb.sheetnames:
                raise RuntimeError("Journals sheet not found. Run the 'Journals' endpoint first to build JournalIDs.")

            journals_cols = ENDPOINTS["Journals"]["columns"]
            ws_journals = wb["Journals"]
            _ensure_sheet_header(ws_journals, journals_cols)

            header_map = _load_journals_header_map(ws_journals, journals_cols)

            proc_cols = ["JournalID", "JournalNumber", "CreatedDateUTC", "HeaderHash", "ProcessedAtUTC", "LineCount"]
            ws_proc = _ensure_sheet(wb, JOURNALLINES_PROCESSED_SHEET, proc_cols)

            processed_hash: Dict[str, str] = {}
            if ws_proc.max_row >= 2:
                header_row = [c.value for c in ws_proc[1]]
                idx_jid = header_row.index("JournalID") + 1
                idx_hh = header_row.index("HeaderHash") + 1
                for r in range(ws_proc.max_row, 1, -1):
                    jid = ws_proc.cell(row=r, column=idx_jid).value
                    if jid is None:
                        continue
                    sjid = str(jid)
                    if sjid in processed_hash:
                        continue
                    hh = ws_proc.cell(row=r, column=idx_hh).value
                    processed_hash[sjid] = "" if hh is None else str(hh)

            to_process: List[str] = []
            changed_ids: Set[str] = set()
            for jid, meta in header_map.items():
                cur_h = meta.get("HeaderHash") or ""
                prev_h = processed_hash.get(jid)
                if prev_h is None:
                    to_process.append(jid)
                elif prev_h != cur_h:
                    to_process.append(jid)
                    changed_ids.add(jid)

            if changed_ids and sheet_name in wb.sheetnames:
                _rewrite_journallines_sheet_excluding(wb, sheet_name, selected_columns, changed_ids)

            if ck and ck.get("endpoint") == "JournalLines" and ck.get("excel_path") == excel_path:
                pending_ids = [str(x) for x in (ck.get("pending_ids") or [])]
                start_index = int(ck.get("journal_index", 0))
                already_written = int(ck.get("rows_written", 0))
                logger.warning(
                    "Resuming JournalLines from checkpoint index=%s rows_written=%s pending=%s",
                    start_index, already_written, len(pending_ids)
                )
            else:
                pending_ids = list(sorted(to_process))
                start_index = 0
                already_written = 0
                _save_checkpoint(checkpoint_path, {
                    "endpoint": "JournalLines",
                    "excel_path": excel_path,
                    "started_at": utc_now_iso(),
                    "pending_ids": pending_ids,
                    "journal_index": 0,
                    "rows_written": 0,
                    "mode": "FILL_MISSING_OR_CHANGED_FROM_JOURNALS",
                })

            if not pending_ids or start_index >= len(pending_ids):
                mode = "FILL_MISSING_OR_CHANGED_FROM_JOURNALS"
                append_run_log(wb, endpoint_name, mode, 0, "OK", None)
                _delete_checkpoint(checkpoint_path)
                save_workbook_atomic(wb, excel_path)
                return 0, "OK", mode, None

            if sheet_name not in wb.sheetnames:
                ws_lines = wb.create_sheet(sheet_name)
                ws_lines.append(selected_columns)
            else:
                ws_lines = wb[sheet_name]
                _ensure_sheet_header(ws_lines, selected_columns)

            existing_keys = _load_existing_row_keys(ws_lines, selected_columns)

            buffer: List[Dict[str, Any]] = []
            rows_written = already_written
            mode = "FILL_MISSING_OR_CHANGED_FROM_JOURNALS"
            proc_appended = 0

            def _append_rows_local(row_dicts: List[Dict[str, Any]]) -> int:
                count = 0
                for rd in row_dicts:
                    values = [get_by_path(rd, col) for col in selected_columns]
                    k = _row_key_from_values(values)
                    if k in existing_keys:
                        continue
                    existing_keys.add(k)
                    ws_lines.append(values)
                    count += 1
                return count

            def _get_json_detail(journal_id: str) -> Optional[Dict[str, Any]]:
                detail_url = f"{XERO_API_BASE}/Journals/{journal_id}"
                r = _get_json(detail_url, headers=headers)
                if r.status_code == 304:
                    return None
                jd = r.json()
                if isinstance(jd, dict):
                    if "Journals" in jd and jd["Journals"]:
                        return jd["Journals"][0]
                    if "Journal" in jd:
                        return jd["Journal"]
                return None

            def _detail_to_rows(journal_obj: Dict[str, Any]) -> List[Dict[str, Any]]:
                out = []
                source_type = journal_obj.get("SourceType")
                source_id = journal_obj.get("SourceID")
                for line in (journal_obj.get("JournalLines", []) or []):
                    line2 = dict(line) if isinstance(line, dict) else {"_raw": line}
                    if line2.get("LineAmount") in (None, ""):
                        fallback_amt = line2.get("NetAmount")
                        if fallback_amt in (None, ""):
                            fallback_amt = line2.get("GrossAmount")
                        if fallback_amt in (None, ""):
                            fallback_amt = line2.get("Amount")
                        if fallback_amt not in (None, ""):
                            line2["LineAmount"] = fallback_amt
                    out.append({
                        "JournalID": journal_obj.get("JournalID"),
                        "JournalNumber": journal_obj.get("JournalNumber"),
                        "JournalDate": journal_obj.get("JournalDate"),
                        "SourceType": source_type,
                        "SourceID": source_id,
                        "JournalLine": line2,
                    })
                return out

            for idx in range(start_index, len(pending_ids)):
                jid = pending_ids[idx]
                meta = header_map.get(jid, {})
                jnum = meta.get("JournalNumber")
                cdu = meta.get("CreatedDateUTC")
                hhash = meta.get("HeaderHash") or ""

                jobj = _get_json_detail(jid)
                if not jobj:
                    ws_proc.append([jid, jnum, cdu, hhash, utc_now_iso(), 0])
                    proc_appended += 1
                else:
                    rows = _detail_to_rows(jobj)
                    buffer.extend(rows)
                    ws_proc.append([jid, jnum, cdu, hhash, utc_now_iso(), len(rows)])
                    proc_appended += 1

                if len(buffer) >= JOURNALLINES_FLUSH_EVERY:
                    rows_written += _append_rows_local(buffer)
                    buffer = []
                    autosize_columns(ws_lines)
                    autosize_columns(ws_proc)
                    save_workbook_atomic(wb, excel_path)
                    _save_checkpoint(checkpoint_path, {
                        "endpoint": "JournalLines",
                        "excel_path": excel_path,
                        "started_at": (ck.get("started_at") if ck else utc_now_iso()),
                        "pending_ids": pending_ids,
                        "journal_index": idx + 1,
                        "rows_written": rows_written,
                        "mode": mode,
                    })

            if buffer:
                rows_written += _append_rows_local(buffer)
                autosize_columns(ws_lines)

            autosize_columns(ws_proc)
            save_workbook_atomic(wb, excel_path)

            _delete_checkpoint(checkpoint_path)
            append_run_log(wb, endpoint_name, mode, rows_written, "OK", None)
            save_workbook_atomic(wb, excel_path)

            logger.info(
                "JournalLines fill complete: lines_written=%s journals_touched=%s changed=%s markers_appended=%s",
                rows_written, len(pending_ids), len(changed_ids), proc_appended
            )
            return rows_written, "OK", mode, None

        # -----------------------
        # Generic: Organisation, Contacts, Invoices, Payments, etc.
        # -----------------------
        if endpoint_name in ENDPOINTS and not ENDPOINTS[endpoint_name].get("special"):
            items = _fetch_generic_endpoint(endpoint_name, headers, incremental_since_iso)
            sheet_name = excel_safe_sheet_name(endpoint_name)
            cols = selected_columns
            mode = f"INCREMENTAL since {incremental_since_iso}" if incremental_since_iso else "FULL"
            if incremental_since_iso and not items:
                rows_written = 0
                append_run_log(wb, endpoint_name, mode, 0, "OK", None)
                save_workbook_atomic(wb, excel_path)
                return 0, "OK", mode, None
            if incremental_since_iso and items:
                key_col = cols[0] if cols else endpoint_columns(endpoint_name)[0]
                rows_written = merge_sheet_selected_columns(wb, sheet_name, items, cols, key_col)
            else:
                rows_written = write_sheet_selected_columns(wb, sheet_name, items, cols)
            append_run_log(wb, endpoint_name, mode, rows_written, "OK", None)
            save_workbook_atomic(wb, excel_path)
            logger.info("Generic endpoint %s: wrote %s rows (mode=%s)", endpoint_name, rows_written, mode)
            return rows_written, "OK", mode, None

        raise ValueError(f"Unknown endpoint: {endpoint_name}")

    except Exception as e:
        mode = "FULL" if not incremental_since_iso else f"INCREMENTAL since {incremental_since_iso}"
        append_run_log(wb, endpoint_name, mode, 0, "FAILED", str(e))
        backup_dir = os.path.join(os.path.dirname(excel_path), "backups")
        backup_file(excel_path, backup_dir, f"excel_before_{excel_safe_sheet_name(endpoint_name)}_FAILED")
        save_workbook_atomic(wb, excel_path)
        logger.exception("Run failed for %s (mode=%s)", endpoint_name, mode)
        return 0, "FAILED", mode, str(e)